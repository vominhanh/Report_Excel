import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import os
from datetime import datetime
import matplotlib
matplotlib.use('Agg')  # Sử dụng backend không GUI cho Flask
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.ticker import FuncFormatter
import io
from PIL import Image as PILImage

class ExcelProcessor:
    def __init__(self, file_paths):
        self.file_paths = file_paths
        self.data = {}
        
    def process(self):
        """Xử lý các file Excel và trích xuất dữ liệu theo SKU"""
        try:
            all_skus = set()
            
            # Đọc tất cả các file
            for file_path in self.file_paths:
                df_dict = pd.read_excel(file_path, sheet_name=None)
                
                # Tìm sheet Performance hoặc sheet chứa danh sách SKU
                performance_sheet = None
                for sheet_name in df_dict.keys():
                    if 'performance' in sheet_name.lower() or 'tổng' in sheet_name.lower():
                        performance_sheet = df_dict[sheet_name]
                        break
                
                if performance_sheet is None:
                    # Nếu không tìm thấy, lấy sheet đầu tiên
                    performance_sheet = list(df_dict.values())[0]
                
                # Tìm cột chứa mã sản phẩm/SKU/ASIN
                sku_column = None
                for col in performance_sheet.columns:
                    col_lower = str(col).lower()
                    if any(keyword in col_lower for keyword in ['sku', 'asin', 'sản phẩm', 'mã']):
                        sku_column = col
                        break
                
                if sku_column is None:
                    # Lấy cột thứ 2 (thường là ASIN)
                    sku_column = performance_sheet.columns[1] if len(performance_sheet.columns) > 1 else performance_sheet.columns[0]
                
                # Lấy danh sách SKU
                skus = performance_sheet[sku_column].dropna().unique()
                all_skus.update(skus)
                
                # Xử lý từng SKU
                for sku in skus:
                    if sku not in self.data:
                        self.data[sku] = {
                            'product_name': None,
                            '2024': pd.DataFrame(),
                            '2025': pd.DataFrame()
                        }
                    
                    # Lấy tên sản phẩm
                    product_col = None
                    for col in performance_sheet.columns:
                        if 'sản phẩm' in str(col).lower() or 'product' in str(col).lower():
                            product_col = col
                            break
                    
                    if product_col:
                        product_row = performance_sheet[performance_sheet[sku_column] == sku]
                        if not product_row.empty:
                            self.data[sku]['product_name'] = product_row.iloc[0][product_col]
                    
                    # Tìm và lấy dữ liệu từ sheet 2024 và 2025
                    for year in ['2024', '2025']:
                        if year in df_dict:
                            year_df = df_dict[year]
                            # Tìm dữ liệu của SKU trong sheet năm
                            sku_data = year_df[year_df.apply(lambda row: row.astype(str).str.contains(str(sku), case=False).any(), axis=1)]
                            
                            if not sku_data.empty:
                                if self.data[sku][year].empty:
                                    self.data[sku][year] = sku_data
                                else:
                                    self.data[sku][year] = pd.concat([self.data[sku][year], sku_data], ignore_index=True)
            
            # Lấy tất cả dữ liệu SKU (có thể chỉ có 2024 hoặc chỉ có 2025)
            filtered_data = {}
            filtered_skus = []
            
            for sku, sku_data in self.data.items():
                has_2024 = not sku_data['2024'].empty
                has_2025 = not sku_data['2025'].empty
                
                # Giữ lại SKU nếu có ít nhất 1 năm có dữ liệu
                if has_2024 or has_2025:
                    filtered_data[sku] = sku_data
                    filtered_skus.append(sku)
            
            return {
                'success': True,
                'skus': filtered_skus,
                'data': filtered_data
            }
        
        except Exception as e:
            return {'error': f'Lỗi xử lý file Excel: {str(e)}'}
    
    def create_output_excel(self, data):
        """Tạo file Excel output với sheet riêng cho mỗi SKU"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f'analysis_report_{timestamp}.xlsx'
            output_path = os.path.join('outputs', output_filename)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Tạo dữ liệu cho sheet so sánh 2024 vs 2025
                comparison_data = []
                summary_data_2024 = []
                summary_data_2025 = []
                
                for sku, sku_data in data.items():
                    product_name = sku_data.get('product_name', 'N/A')
                    
                    # Dữ liệu 2024
                    quantity_2024 = revenue_2024 = ad_spent_2024 = 0
                    if not sku_data['2024'].empty:
                        quantity_col = self._find_column(sku_data['2024'], ['số lượng bán ra', 'quantity', 'units sold', 'sold'])
                        revenue_col = self._find_column(sku_data['2024'], ['doanh số', 'revenue', 'tổng doanh', 'sales'])
                        ad_cost_col = self._find_column(sku_data['2024'], ['chi phí quảng cáo', 'ad cost', 'advertising', 'quảng cáo', 'ad spent'])
                        
                        if quantity_col:
                            quantity_2024 = pd.to_numeric(sku_data['2024'][quantity_col], errors='coerce').fillna(0).sum()
                        if revenue_col:
                            revenue_2024 = pd.to_numeric(sku_data['2024'][revenue_col], errors='coerce').fillna(0).sum()
                        if ad_cost_col:
                            ad_spent_2024 = pd.to_numeric(sku_data['2024'][ad_cost_col], errors='coerce').fillna(0).sum()
                    
                    # Dữ liệu 2025
                    quantity_2025 = revenue_2025 = ad_spent_2025 = 0
                    if not sku_data['2025'].empty:
                        quantity_col = self._find_column(sku_data['2025'], ['số lượng bán ra', 'quantity', 'units sold', 'sold'])
                        revenue_col = self._find_column(sku_data['2025'], ['doanh số', 'revenue', 'tổng doanh', 'sales'])
                        ad_cost_col = self._find_column(sku_data['2025'], ['chi phí quảng cáo', 'ad cost', 'advertising', 'quảng cáo', 'ad spent'])
                        
                        if quantity_col:
                            quantity_2025 = pd.to_numeric(sku_data['2025'][quantity_col], errors='coerce').fillna(0).sum()
                        if revenue_col:
                            revenue_2025 = pd.to_numeric(sku_data['2025'][revenue_col], errors='coerce').fillna(0).sum()
                        if ad_cost_col:
                            ad_spent_2025 = pd.to_numeric(sku_data['2025'][ad_cost_col], errors='coerce').fillna(0).sum()
                    
                    # Tính TACOS cho từng năm
                    tacos_2024 = (ad_spent_2024 / revenue_2024 * 100) if revenue_2024 > 0 else 0
                    tacos_2025 = (ad_spent_2025 / revenue_2025 * 100) if revenue_2025 > 0 else 0
                    
                    # Thêm vào dữ liệu so sánh
                    comparison_data.append({
                        'Mã SKU': sku,
                        'Sản phẩm': product_name,
                        'Số lượng 2024': int(quantity_2024),
                        'Doanh số 2024': revenue_2024,
                        'Ad spent 2024': ad_spent_2024,
                        'TACOS 2024': f"{tacos_2024:.2f}%",
                        'Số lượng 2025': int(quantity_2025),
                        'Doanh số 2025': revenue_2025,
                        'Ad spent 2025': ad_spent_2025,
                        'TACOS 2025': f"{tacos_2025:.2f}%"
                    })
                    
                    # Thêm vào dữ liệu 2024
                    if quantity_2024 > 0 or revenue_2024 > 0:
                        category_2024 = self._get_category(tacos_2024)
                        summary_data_2024.append({
                            'Mã SKU': sku,
                            'Sản phẩm': product_name,
                            'Số lượng bán ra': int(quantity_2024),
                            'Tổng doanh số': revenue_2024,
                            'Tổng Ad spent': ad_spent_2024,
                            'Tacos': f"{tacos_2024:.2f}%",
                            'Phân loại': category_2024
                        })
                    
                    # Thêm vào dữ liệu 2025
                    if quantity_2025 > 0 or revenue_2025 > 0:
                        category_2025 = self._get_category(tacos_2025)
                        summary_data_2025.append({
                            'Mã SKU': sku,
                            'Sản phẩm': product_name,
                            'Số lượng bán ra': int(quantity_2025),
                            'Tổng doanh số': revenue_2025,
                            'Tổng Ad spent': ad_spent_2025,
                            'Tacos': f"{tacos_2025:.2f}%",
                            'Phân loại': category_2025
                        })
                
                # Tạo sheet so sánh 2024 vs 2025
                comparison_df = pd.DataFrame(comparison_data)
                comparison_df.to_excel(writer, sheet_name='TỔNG PERFORMANCE', index=False)
                
                # Thêm biểu đồ so sánh vào sheet TỔNG PERFORMANCE
                if comparison_df is not None and not comparison_df.empty:
                    self._add_comparison_chart(writer.book, 'TỔNG PERFORMANCE', comparison_df)
                
                # Tạo sheet cho từng SKU + chèn biểu đồ trực tiếp trong Excel
                for sku, sku_data in data.items():
                    # Sử dụng tên sản phẩm làm tên sheet
                    product_name = sku_data.get('product_name', str(sku))
                    if not product_name or product_name == 'N/A' or pd.isna(product_name):
                        product_name = str(sku)
                    
                    # Làm sạch tên sheet (Excel giới hạn 31 ký tự, không chứa ký tự đặc biệt)
                    sheet_name = str(product_name)[:31]
                    # Loại bỏ ký tự không hợp lệ cho tên sheet Excel
                    invalid_chars = ['[', ']', '*', '?', ':', '\\', '/']
                    for char in invalid_chars:
                        sheet_name = sheet_name.replace(char, '')

                    combined_data = []
                    if not sku_data['2024'].empty:
                        df_2024 = sku_data['2024'].copy()
                        df_2024.insert(0, 'Năm', '2024')
                        combined_data.append(df_2024)

                    if not sku_data['2025'].empty:
                        df_2025 = sku_data['2025'].copy()
                        df_2025.insert(0, 'Năm', '2025')
                        combined_data.append(df_2025)

                    if combined_data:
                        combined_df = pd.concat(combined_data, ignore_index=True)
                        
                        # Xử lý cột thời gian: ghép các cột Unnamed thành "Oct - 2nd", "Nov - 1st"...
                        combined_df = self._process_time_columns(combined_df)
                        
                        # Bỏ các cột Unnamed còn lại
                        combined_df = combined_df.loc[:, ~combined_df.columns.str.contains('^Unnamed', na=False)]
                        
                        # Thêm cột "Tacos an toàn" = 30% cho tất cả các dòng
                        if 'Tacos an toàn' not in combined_df.columns:
                            combined_df['Tacos an toàn'] = 0.30
                        
                        combined_df.to_excel(writer, sheet_name=sheet_name, index=False)

                        # Thêm biểu đồ trực tiếp vào sheet SKU
                        product_name = sku_data.get('product_name')
                        self._add_charts_to_sheet(
                            writer.book,
                            sheet_name,
                            combined_df,
                            product_name=product_name,
                            sku=str(sku)
                        )
                
                # Định dạng file Excel
                self._format_excel(writer)
            
            return output_filename
        
        except Exception as e:
            print(f"Lỗi tạo file Excel: {str(e)}")
            return None
    
    def _find_column(self, df, keywords):
        """Tìm cột dựa trên từ khóa"""
        for col in df.columns:
            col_lower = str(col).lower()
            if any(keyword in col_lower for keyword in keywords):
                return col
        return None
    
    def _get_category(self, tacos_percent):
        """Xác định phân loại dựa trên TACOS"""
        if tacos_percent == 0:
            return "Tốt"
        elif tacos_percent <= 30:
            return "Tốt" 
        elif tacos_percent <= 50:
            return "Xấu"
        else:
            return "TB"  # Trung bình
    
    def _process_time_columns(self, df):
        """Xử lý và ghép các cột Unnamed thành cột Thời gian"""
        try:
            # Tìm các cột Unnamed ở đầu (thường là cột 1, 2)
            unnamed_cols = [col for col in df.columns if 'Unnamed' in str(col)]
            
            if len(unnamed_cols) >= 2:
                # Ghép 2 cột đầu tiên thành "Oct - 2nd", "Nov - 1st"...
                first_col = unnamed_cols[0]
                second_col = unnamed_cols[1]
                
                # Tạo cột Thời gian mới với auto-increment cho nan
                time_values = []
                month_day_counter = {}  # Đếm ngày cho mỗi tháng khi gặp nan
                
                for i in range(len(df)):
                    first_val = str(df.iloc[i][first_col])
                    second_val = str(df.iloc[i][second_col])
                    
                    if 'nan' in first_val and 'nan' in second_val:
                        time_values.append('')
                    elif 'nan' in second_val:
                        # Trường hợp "Oct - nan"
                        month_name = first_val.lower()
                        
                        # Tự động tăng ngày cho tháng này
                        if month_name not in month_day_counter:
                            month_day_counter[month_name] = 1
                        else:
                            month_day_counter[month_name] += 1
                        
                        # Tạo nhãn với ngày tự động
                        if month_day_counter[month_name] == 1:
                            suffix = 'st'
                        elif month_day_counter[month_name] == 2:
                            suffix = 'nd'
                        elif month_day_counter[month_name] == 3:
                            suffix = 'rd'
                        else:
                            suffix = 'th'
                        
                        time_val = f"{first_val} - {month_day_counter[month_name]}{suffix}"
                        time_values.append(time_val)
                    else:
                        # Trường hợp bình thường
                        time_val = first_val + ' - ' + second_val
                        time_values.append(time_val)
                
                df['Thời gian'] = time_values
                
                # Làm sạch giá trị (bỏ nan, None...)
                df['Thời gian'] = df['Thời gian'].replace('nan - nan', '')
                df['Thời gian'] = df['Thời gian'].replace('None - None', '')
                
                # Di chuyển cột Thời gian lên đầu (sau cột Năm nếu có)
                cols = df.columns.tolist()
                cols.remove('Thời gian')
                
                # Tìm vị trí chèn (sau cột Năm nếu có, không thì đầu tiên)
                if 'Năm' in cols:
                    insert_pos = cols.index('Năm') + 1
                else:
                    insert_pos = 0
                
                cols.insert(insert_pos, 'Thời gian')
                df = df[cols]
            
            return df
        
        except Exception as e:
            print(f"Lỗi xử lý cột thời gian: {str(e)}")
            return df

    def _style_chart_title(self, chart, title_text, color="C00000", size=1400):
        """Đổi màu/độ đậm tiêu đề chart - phiên bản đơn giản"""
        # Phiên bản openpyxl cũ không hỗ trợ RichText
        # Chỉ set title text đơn giản
        pass
    
    def _format_excel(self, writer):
        """Định dạng file Excel"""
        try:
            workbook = writer.book
            
            # Định dạng cho tất cả các sheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Định dạng header
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=11)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Tự động điều chỉnh độ rộng cột
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    header_value = str(column[0].value).lower() if column[0].value else ""
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                    # Định dạng số theo cột
                    if 'tacos an toàn' in header_value:
                        for cell in column[1:]:
                            cell.number_format = '0.00%'
                    elif 'tacos' in header_value:
                        for cell in column[1:]:
                            cell.number_format = '0.00%'
                    elif any(keyword in header_value for keyword in ['tổng doanh', 'chi phí']):
                        for cell in column[1:]:
                            cell.number_format = '"$"#,##0.00'
        
        except Exception as e:
            print(f"Lỗi định dạng Excel: {str(e)}")

    def _add_charts_to_sheet(self, workbook, sheet_name, df, product_name=None, sku=None):
        """Tạo biểu đồ bằng matplotlib và chèn vào Excel như hình ảnh."""
        try:
            ws = workbook[sheet_name]

            if df is None or df.empty:
                print(f"Sheet {sheet_name}: DataFrame rỗng, bỏ qua biểu đồ")
                return

            year_col = self._find_column(df, ['năm', 'year'])
            time_col = self._find_column(df, ['thời gian', 'time', 'ngày', 'date', 'tuần', 'week'])
            if time_col is None and len(df.columns) > 1:
                time_col = df.columns[1]

            revenue_col = self._find_column(df, ['tổng doanh số', 'doanh số', 'revenue', 'sales'])
            ad_cost_col = self._find_column(df, ['chi phí quảng cáo', 'ad cost', 'advertising', 'quảng cáo'])
            tacos_col = self._find_column(df, ['tacos'])
            safe_tacos_col = self._find_column(df, ['tacos an toàn', 'tacos an toan', 'safe tacos'])

            print(f"Sheet {sheet_name}: year_col={year_col}, time_col={time_col}, revenue_col={revenue_col}")

            if not year_col or not revenue_col or not time_col:
                print(f"Sheet {sheet_name}: Thiếu cột quan trọng, bỏ qua biểu đồ")
                return

            def _get_display_name():
                name = product_name
                if name is None or (isinstance(name, float) and pd.isna(name)):
                    name = sku
                if name is None:
                    name = sheet_name
                name = str(name).strip()
                if not name:
                    name = sheet_name
                return name.upper()

            display_name = _get_display_name()

            # Lấy danh sách năm
            years = [str(y) for y in sorted(df[year_col].dropna().unique())]
            if not years:
                return

            # Tính toán vị trí chèn biểu đồ
            max_data_col = len(df.columns)
            start_chart_col_idx = max_data_col + 2
            start_chart_col = get_column_letter(start_chart_col_idx)
            second_chart_col = get_column_letter(start_chart_col_idx + 10)

            for i, year in enumerate(years):
                year_df = df[df[year_col].astype(str) == year].copy()
                if year_df.empty:
                    continue

                # Lọc và sắp xếp dữ liệu
                # Loại bỏ các dòng có doanh số = 0 hoặc NaN
                year_df = year_df.copy()
                year_df['revenue_numeric'] = pd.to_numeric(year_df[revenue_col], errors='coerce')
                
                # Nếu có cột chi phí quảng cáo, lọc cả 2 cột
                if ad_cost_col:
                    year_df['ad_cost_numeric'] = pd.to_numeric(year_df[ad_cost_col], errors='coerce')
                    # Loại bỏ dòng nào có cả doanh số và chi phí = 0 hoặc NaN
                    year_df = year_df[
                        (year_df['revenue_numeric'].notna() & (year_df['revenue_numeric'] > 0)) |
                        (year_df['ad_cost_numeric'].notna() & (year_df['ad_cost_numeric'] > 0))
                    ]
                else:
                    # Chỉ lọc doanh số > 0
                    year_df = year_df[
                        (year_df['revenue_numeric'].notna()) & (year_df['revenue_numeric'] > 0)
                    ]
                
                # Nếu không còn dữ liệu sau khi lọc, bỏ qua
                if year_df.empty:
                    print(f"Không có dữ liệu hợp lệ cho năm {year}, bỏ qua biểu đồ")
                    continue
                
                # Sắp xếp theo thời gian tăng dần (dựa trên cột thời gian)
                # Tạo key sắp xếp từ cột thời gian
                def create_sort_key(time_str):
                    try:
                        time_str = str(time_str).lower()
                        if 'jan' in time_str: month = 1
                        elif 'feb' in time_str: month = 2
                        elif 'mar' in time_str: month = 3
                        elif 'apr' in time_str: month = 4
                        elif 'may' in time_str: month = 5
                        elif 'jun' in time_str: month = 6
                        elif 'jul' in time_str: month = 7
                        elif 'aug' in time_str: month = 8
                        elif 'sep' in time_str: month = 9
                        elif 'oct' in time_str: month = 10
                        elif 'nov' in time_str: month = 11
                        elif 'dec' in time_str: month = 12
                        else: month = 0
                        
                        # Tìm số ngày
                        import re
                        day_match = re.search(r'(\d+)', time_str)
                        day = int(day_match.group(1)) if day_match else 1  # Mặc định ngày 1 nếu không tìm thấy
                        
                        return (month, day)
                    except:
                        return (0, 0)
                
                year_df['sort_key'] = year_df[time_col].apply(create_sort_key)
                year_df = year_df.sort_values('sort_key')

                # ========== Biểu đồ 1: Doanh số + Chi phí quảng cáo ==========
                plt.style.use('default')
                fig, ax1 = plt.subplots(figsize=(14, 8))  # Tăng kích thước từ 12x6 lên 14x8
                
                # Dữ liệu cho biểu đồ - rút gọn nhãn thời gian chỉ hiển thị ngày/tháng
                time_labels_raw = year_df[time_col].astype(str).tolist()
                time_labels = []
                month_day_counter = {}  # Đếm ngày cho mỗi tháng
                
                for label in time_labels_raw:
                    if 'nan' in label or 'None' in label or label.strip() == '':
                        time_labels.append('')
                    else:
                        import re
                        # Mapping tháng
                        month_map = {
                            'jan': '1', 'feb': '2', 'mar': '3', 'apr': '4',
                            'may': '5', 'jun': '6', 'jul': '7', 'aug': '8', 
                            'sep': '9', 'oct': '10', 'nov': '11', 'dec': '12'
                        }
                        
                        label_lower = label.lower()
                        day = ''
                        month = ''
                        
                        # Tìm tháng
                        for month_name, month_num in month_map.items():
                            if month_name in label_lower:
                                month = month_num
                                break
                        
                        # Tìm ngày (số đầu tiên trong chuỗi, bỏ qua nan)
                        day_match = re.search(r'(\d+)', label)
                        if day_match and 'nan' not in label_lower:
                            day = day_match.group(1)
                        else:
                            # Nếu không có ngày cụ thể hoặc có "nan", tự động tăng ngày cho tháng này
                            if month:
                                if month not in month_day_counter:
                                    month_day_counter[month] = 1
                                else:
                                    month_day_counter[month] += 1
                                day = str(month_day_counter[month])
                            else:
                                day = '1'
                        
                        if month:
                            time_labels.append(f'{day}/{month}')
                        else:
                            time_labels.append(label)
                
                revenue_data = year_df['revenue_numeric'].fillna(0).tolist()
                
                # Cột doanh số
                bars = ax1.bar(time_labels, revenue_data, color='#1F4E78', alpha=0.8, label='Tổng doanh số')
                
                # Định dạng trục Y trái (doanh số)
                ax1.set_ylabel('Tổng doanh số ($)', fontweight='bold')
                ax1.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'${x:,.0f}'))
                
                # Thiết lập 5 mốc cho trục Y
                if max(revenue_data) > 0:
                    max_val = max(revenue_data)
                    import math
                    max_rounded = math.ceil(max_val / 500) * 500
                    ax1.set_ylim(0, max_rounded)
                    ax1.set_yticks([i * max_rounded / 4 for i in range(5)])
                
                ax1.grid(True, alpha=0.3)
                
                # Đường chi phí quảng cáo (nếu có)
                if ad_cost_col and 'ad_cost_numeric' in year_df.columns:
                    ax2 = ax1.twinx()
                    ad_cost_data = year_df['ad_cost_numeric'].fillna(0).tolist()
                    line = ax2.plot(time_labels, ad_cost_data, color='#C00000', linewidth=3, marker='o', label='Chi phí quảng cáo')
                    ax2.set_ylabel('Chi phí quảng cáo ($)', fontweight='bold')
                    ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'${x:,.0f}'))
                    
                    # Kết hợp chú thích của cả 2 trục
                    lines1, labels1 = ax1.get_legend_handles_labels()
                    lines2, labels2 = ax2.get_legend_handles_labels()
                    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left', frameon=True, fancybox=True, shadow=True)
                else:
                    # Chỉ có chú thích doanh số
                    ax1.legend(loc='upper left', frameon=True, fancybox=True, shadow=True)
                
                # Tiêu đề và định dạng
                plt.title(f'{display_name} {year}', fontsize=16, fontweight='bold', color='#C00000')
                
                # Xoay nhãn trục X và điều chỉnh khoảng cách - hiển thị tất cả nhãn
                plt.xticks(range(len(time_labels)), time_labels, rotation=45, ha='right', fontsize=9)
                
                # Đảm bảo khoảng cách giữa các nhãn
                ax1.tick_params(axis='x', which='major', pad=5)
                
                plt.tight_layout()
                
                # Lưu biểu đồ doanh số
                img_buffer1 = io.BytesIO()
                plt.savefig(img_buffer1, format='png', dpi=300, bbox_inches='tight')
                plt.close(fig)  # Đóng figure để giải phóng bộ nhớ
                img_buffer1.seek(0)
                
                # Chèn vào Excel
                img1 = Image(img_buffer1)
                img1.width = 560  # Tăng từ 480 pixels
                img1.height = 336  # Tăng từ 288 pixels
                ws.add_image(img1, f'{start_chart_col}{2 + i * 20}')
                
                plt.close()

                # ========== Biểu đồ 2: TACOS ==========
                if tacos_col:
                    fig, ax = plt.subplots(figsize=(14, 8))  # Tăng kích thước
                    
                    # Dữ liệu TACOS
                    tacos_data = year_df[tacos_col].astype(str).str.rstrip('%').astype(float, errors='ignore')
                    tacos_data = pd.to_numeric(tacos_data, errors='coerce').fillna(0)
                    
                    # Chuyển sang decimal nếu dữ liệu > 1 (nghĩa là đang ở dạng %)
                    if tacos_data.max() > 1:
                        tacos_data = tacos_data / 100
                    
                    # Cột TACOS
                    bars = ax.bar(time_labels, tacos_data, color='#1F4E78', alpha=0.8, label='TACOS')
                    
                    # Đường TACOS an toàn 30%
                    if safe_tacos_col:
                        safe_data = [0.30] * len(time_labels)
                        line = ax.plot(time_labels, safe_data, color='#C00000', linewidth=3, label='TACOS an toàn (30%)', linestyle='--')
                    
                    # Định dạng trục Y
                    ax.set_ylabel('TACOS (%)', fontweight='bold')
                    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'{x:.0%}'))
                    ax.set_ylim(0, max(0.6, tacos_data.max() * 1.1))
                    ax.set_yticks([i * 0.1 for i in range(7)])  # 0%, 10%, 20%, ..., 60%
                    ax.grid(True, alpha=0.3)
                    
                    # Tiêu đề và chú thích
                    plt.title(f'TACOS {year}', fontsize=16, fontweight='bold', color='#C00000')
                    
                    # Xoay nhãn và điều chỉnh hiển thị - hiển thị tất cả nhãn
                    plt.xticks(range(len(time_labels)), time_labels, rotation=45, ha='right', fontsize=9)
                    ax.tick_params(axis='x', which='major', pad=5)
                    
                    # Chú thích
                    ax.legend(loc='upper left', frameon=True, fancybox=True, shadow=True)
                    plt.tight_layout()
                    
                    # Lưu biểu đồ TACOS
                    img_buffer2 = io.BytesIO()
                    plt.savefig(img_buffer2, format='png', dpi=300, bbox_inches='tight')
                    plt.close(fig)  # Đóng figure để giải phóng bộ nhớ
                    img_buffer2.seek(0)
                    
                    # Chèn vào Excel
                    img2 = Image(img_buffer2)
                    img2.width = 560  # Tăng kích thước
                    img2.height = 336
                    ws.add_image(img2, f'{second_chart_col}{2 + i * 20}')
                    
                    plt.close()  # Đóng tất cả figures còn lại

                print(f"Đã thêm biểu đồ matplotlib {year} vào {start_chart_col}{2 + i * 20} và {second_chart_col}{2 + i * 20}")

        except Exception as e:
            print(f"Lỗi tạo biểu đồ matplotlib: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _add_comparison_chart(self, workbook, sheet_name, df):
        """Tạo biểu đồ so sánh giữa 2024 và 2025"""
        try:
            ws = workbook[sheet_name]
            
            # Tính tổng các chỉ số cho 2024 và 2025
            total_quantity_2024 = df['Số lượng 2024'].sum()
            total_revenue_2024 = df['Doanh số 2024'].sum()
            total_ad_spent_2024 = df['Ad spent 2024'].sum()
            
            total_quantity_2025 = df['Số lượng 2025'].sum()
            total_revenue_2025 = df['Doanh số 2025'].sum()
            total_ad_spent_2025 = df['Ad spent 2025'].sum()
            
            # Tính TACOS tổng
            tacos_2024 = (total_ad_spent_2024 / total_revenue_2024 * 100) if total_revenue_2024 > 0 else 0
            tacos_2025 = (total_ad_spent_2025 / total_revenue_2025 * 100) if total_revenue_2025 > 0 else 0
            
            # Tạo bảng tổng hợp
            summary_data = [
                ['Năm', 'Số lượng bán ra', 'Tổng doanh số', 'Tổng Ad spent', 'Tacos'],
                ['2025', total_quantity_2025, total_revenue_2025, total_ad_spent_2025, f"{tacos_2025:.2f}%"],
                ['2024', total_quantity_2024, total_revenue_2024, total_ad_spent_2024, f"{tacos_2024:.2f}%"]
            ]
            
            # Chèn bảng vào Excel (bắt đầu từ cột sau dữ liệu chính)
            start_row = len(df) + 5
            for i, row in enumerate(summary_data):
                for j, value in enumerate(row):
                    cell = ws.cell(row=start_row + i, column=j + 1, value=value)
                    if i == 0:  # Header
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF')
            
            # Tạo biểu đồ so sánh
            plt.style.use('default')
            fig, ax = plt.subplots(figsize=(12, 8))
            
            # Dữ liệu cho biểu đồ
            years = ['2025', '2024']
            quantities = [total_quantity_2025, total_quantity_2024]
            revenues = [total_revenue_2025, total_revenue_2024]
            ad_spents = [total_ad_spent_2025, total_ad_spent_2024]
            
            x = range(len(years))
            width = 0.25
            
            # Tạo các cột
            bars1 = ax.bar([i - width for i in x], quantities, width, label='Số lượng bán ra', color='#1F4E78', alpha=0.8)
            bars2 = ax.bar(x, revenues, width, label='Tổng doanh số', color='#C00000', alpha=0.8)
            bars3 = ax.bar([i + width for i in x], ad_spents, width, label='Tổng Ad spent', color='#FFC000', alpha=0.8)
            
            # Định dạng biểu đồ
            ax.set_xlabel('Năm', fontweight='bold')
            ax.set_ylabel('Giá trị', fontweight='bold')
            ax.set_title('Số lượng bán ra and Tổng doanh số', fontsize=14, fontweight='bold')
            ax.set_xticks(x)
            ax.set_xticklabels(years)
            ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1))
            ax.grid(True, alpha=0.3)
            
            # Định dạng trục Y với dấu phẩy
            ax.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f'{x:,.0f}'))
            
            plt.tight_layout()
            
            # Lưu biểu đồ
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
            plt.close(fig)
            img_buffer.seek(0)
            
            # Chèn vào Excel
            img = Image(img_buffer)
            img.width = 600
            img.height = 400
            ws.add_image(img, f'A{start_row + len(summary_data) + 2}')
            
            print(f"Đã thêm biểu đồ so sánh vào sheet {sheet_name}")
            
        except Exception as e:
            print(f"Lỗi tạo biểu đồ so sánh: {str(e)}")
            import traceback
            traceback.print_exc()



